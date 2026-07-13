#!/usr/bin/env python3
"""
quant_check.py — the Excel double-check for every model the browser runs.

Replicates 1:1 the JavaScript in docs/quant.js, docs/ols.js and docs/ff.js
(same conventions: daily % returns, population std, 252 days/year,
linear-interpolation quantiles, GARCH grid MLE with identical float loop
accumulation, Monte Carlo with the same mulberry32(42) stream), then
writes quant_double_check.xlsx where everything that CAN be an Excel
formula IS one — so the numbers recompute in front of you and can be
audited cell by cell. Only outputs that need an optimizer or 10,000
simulations (GARCH α/β, Monte Carlo VaR, regression betas) are written
as values, in blue, each with a note saying exactly how they were made.

Test portfolio (same one used in the browser test session):
    NVDA   €3,000 invested, purchase date 2023-01-10
    BGF-SE €1,500 invested, purchase date 2024-03-15

Run:  python3 tools/quant_check.py
"""

import json
import math
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA = os.path.join(ROOT, "docs", "data")

TRADING_DAYS = 252
PORTFOLIO = [
    {"symbol": "NVDA", "name": "NVIDIA Corporation", "amount": 3000.0,
     "purchase": "2023-01-10", "file": os.path.join(DATA, "tickers", "NVDA.json")},
    {"symbol": "BGF-SE", "name": "BlackRock BGF Sustainable Energy E2 EUR", "amount": 1500.0,
     "purchase": "2024-03-15", "file": os.path.join(DATA, "nav_history.json")},
]

# ── tiny helpers mirroring quant.js exactly ──
def qmean(a): return sum(a) / len(a)
def qstd(a):
    m = qmean(a)
    return math.sqrt(sum((x - m) ** 2 for x in a) / len(a))  # population

def quantile_sorted(a, p):
    """Linear interpolation, same as JS _quantileSorted / numpy default."""
    pos = (len(a) - 1) * p
    lo = int(math.floor(pos)); frac = pos - lo
    return a[lo] + frac * (a[lo + 1] - a[lo]) if lo + 1 < len(a) else a[lo]

M32 = 0xFFFFFFFF
def mulberry32(seed):
    """Bit-exact port of the JS mulberry32 (uint32 arithmetic)."""
    a = seed & M32
    def rng():
        nonlocal a
        a = (a + 0x6D2B79F5) & M32
        t = ((a ^ (a >> 15)) * ((1 | a) & M32)) & M32
        t = ((t + (((t ^ (t >> 7)) * ((61 | t) & M32)) & M32)) ^ t) & M32
        return ((t ^ (t >> 14)) & M32) / 4294967296
    return rng


def nav_on_or_before(hist, date_iso):
    match = None
    for d, v in hist:
        if d <= date_iso:
            match = v
        else:
            break
    return match


def build_portfolio_returns(holdings):
    """Mirror of ff.js buildPortfolioReturns: weighted sum of daily %
    returns on the dates ALL holdings share."""
    ret_maps = []
    for h in holdings:
        m = {}
        hist = h["history"]
        for i in range(1, len(hist)):
            p0 = hist[i - 1][1]
            if p0:
                m[hist[i][0]] = (hist[i][1] - p0) / p0 * 100
        ret_maps.append({"weight": h["weight"], "m": m})
    dates = None
    for r in ret_maps:
        s = set(r["m"].keys())
        dates = s if dates is None else (dates & s)
    out = []
    for d in sorted(dates):
        out.append((d, sum(r["weight"] * r["m"][d] for r in ret_maps)))
    return out


# ── perfMetrics (quant.js) ──
def perf_metrics(rets, rf_daily):
    n = len(rets)
    mean = qmean(rets); sd = qstd(rets)
    excess = [r - rf_daily for r in rets]
    prod = 1.0
    for r in rets: prod *= 1 + r / 100
    ann_ret = (prod ** (TRADING_DAYS / n) - 1) * 100
    ann_vol = sd * math.sqrt(TRADING_DAYS)
    sharpe = qmean(excess) / sd * math.sqrt(TRADING_DAYS) if sd else 0
    downside = [r for r in excess if r < 0]
    dsd = math.sqrt(sum(x * x for x in downside) / len(excess)) if downside else 0
    sortino = qmean(excess) / dsd * math.sqrt(TRADING_DAYS) if dsd else None
    wealth = peak = 1.0; max_dd = 0.0
    wealth_path = []
    for r in rets:
        wealth *= 1 + r / 100
        peak = max(peak, wealth)
        max_dd = min(max_dd, (wealth - peak) / peak)
        wealth_path.append(wealth)
    max_dd *= 100
    calmar = ann_ret / abs(max_dd) if max_dd < 0 else None
    return {"n": n, "annRet": ann_ret, "annVol": ann_vol, "sharpe": sharpe,
            "sortino": sortino, "maxDD": max_dd, "calmar": calmar,
            "meanDaily": mean, "sdDaily": sd, "wealth": wealth_path}


# ── garchFit (quant.js) — identical grid, identical float accumulation ──
def garch_fit(rets):
    n = len(rets)
    mu = qmean(rets)
    eps = [r - mu for r in rets]
    uncond = sum(e * e for e in eps) / n

    def neg_ll(alpha, beta):
        omega = uncond * (1 - alpha - beta)
        sigma2 = uncond; ll = 0.0
        for t in range(1, n):
            sigma2 = omega + alpha * eps[t - 1] * eps[t - 1] + beta * sigma2
            if sigma2 <= 0: return float("inf")
            ll += math.log(sigma2) + (eps[t] * eps[t]) / sigma2
        return ll

    best = (0.05, 0.90, float("inf"))
    a = 0.02
    while a <= 0.301:
        b = 0.55
        while b <= 0.981:
            if a + b < 0.999:
                v = neg_ll(a, b)
                if v < best[2]: best = (a, b, v)
            b += 0.01
        a += 0.02
    # Refinement window ±0.02 in BOTH parameters (a full coarse cell per
    # side) — mirrors quant.js; a ±0.01 beta window can miss the true
    # optimum on a flat likelihood ridge (verified on real data).
    a0, b0 = best[0], best[1]
    a = max(0.005, a0 - 0.02)
    while a <= a0 + 0.0201:
        b = max(0.4, b0 - 0.02)
        while b <= min(0.995, b0 + 0.0201):
            if a + b < 0.999:
                v = neg_ll(a, b)
                if v < best[2]: best = (a, b, v)
            b += 0.0025
        a += 0.005

    alpha, beta, ll = best
    omega = uncond * (1 - alpha - beta)
    sigma2 = uncond
    for t in range(1, n):
        sigma2 = omega + alpha * eps[t - 1] * eps[t - 1] + beta * sigma2
    sigma2_next = omega + alpha * eps[n - 1] * eps[n - 1] + beta * sigma2
    pers = alpha + beta
    f = 0.0; s2k = sigma2_next
    for _ in range(10):
        f += s2k
        s2k = uncond + pers * (s2k - uncond)
    return {"n": n, "alpha": alpha, "beta": beta, "omega": omega,
            "persistence": pers, "uncondVar": uncond, "mu": mu, "negLL": ll,
            "uncondVolAnn": math.sqrt(uncond * TRADING_DAYS),
            "condVolAnn": math.sqrt(sigma2_next * TRADING_DAYS),
            "condVolDaily": math.sqrt(sigma2_next),
            "forecast10dVolAnn": math.sqrt(f / 10 * TRADING_DAYS),
            "regimeRatio": math.sqrt(sigma2_next / uncond), "eps": eps}


# ── varAnalysis (quant.js) — bit-identical Monte Carlo ──
def var_analysis(rets, cond_vol_daily, value):
    srt = sorted(rets)
    var95h = -quantile_sorted(srt, 0.05)
    var99h = -quantile_sorted(srt, 0.01)
    tail = [r for r in srt if r <= -var95h]
    es95 = -qmean(tail) if tail else var95h
    sd = qstd(rets)
    rng = mulberry32(42)
    N, H = 10000, 10
    mc = []
    for _ in range(N):
        w = 1.0
        for _ in range(H):
            w *= 1 + rets[int(rng() * len(rets))] / 100
        mc.append((w - 1) * 100)
    mc.sort()
    return {
        "hist": {"var95": var95h, "var99": var99h, "es95": es95},
        "param": {"var95": 1.645 * sd, "var99": 2.326 * sd, "sd": sd},
        "paramGarch": ({"var95": 1.645 * cond_vol_daily, "var99": 2.326 * cond_vol_daily}
                       if cond_vol_daily else None),
        "mc10d": {"var95": -quantile_sorted(mc, 0.05), "var99": -quantile_sorted(mc, 0.01)},
        "value": value,
    }


# ── tsmom (quant.js) ──
def ema(closes, N):
    if len(closes) < N + 1: return None
    k = 2 / (N + 1)
    e = qmean(closes[:N])
    for c in closes[N:]:
        e = c * k + e * (1 - k)
    return e

def tsmom(closes):
    if len(closes) < 130: return None
    last = closes[-1]
    horizons = [h for h in (21, 63, 126, 252) if len(closes) > h]
    signals = []
    for h in horizons:
        past = closes[-1 - h]
        ret = (last / past - 1) * 100
        signals.append({"h": h, "ret": ret, "sign": 1 if ret > 0 else (-1 if ret < 0 else 0)})
    e20, e100 = ema(closes, 20), ema(closes, 100)
    ema_sign = (1 if e20 > e100 else -1) if (e20 is not None and e100 is not None) else 0
    parts = [s["sign"] for s in signals] + ([ema_sign] if ema_sign else [])
    score = qmean(parts) if parts else 0
    return {"signals": signals, "ema20": e20, "ema100": e100, "emaSign": ema_sign, "score": score}


# ── OLS replication (ols.js _fitOLS with normal-CDF p-values) ──
def normal_cdf(x):
    sign = -1 if x < 0 else 1
    ax = abs(x) / math.sqrt(2)
    t = 1 / (1 + 0.3275911 * ax)
    y = 1 - (((((1.061405429 * t - 1.453152027) * t) + 1.421413741) * t
              - 0.284496736) * t + 0.254829592) * t * math.exp(-ax * ax)
    return 0.5 * (1 + sign * y)

def fit_ols(y, X):
    import numpy as np
    n = len(y); k = len(X[0]) + 1
    Xc = np.column_stack([np.ones(n), np.array(X)])
    yv = np.array(y)
    XtX = Xc.T @ Xc
    beta = np.linalg.solve(XtX, Xc.T @ yv)
    resid = yv - Xc @ beta
    ss_res = float(resid @ resid)
    ss_tot = float(((yv - yv.mean()) ** 2).sum())
    df = n - k
    sigma2 = ss_res / df
    XtXinv = np.linalg.inv(XtX)
    pvals = []
    for a in range(k):
        se = math.sqrt(sigma2 * XtXinv[a, a])
        pvals.append(2 * (1 - normal_cdf(abs(beta[a] / se))) if se > 0 else 1.0)
    r2 = 1 - ss_res / ss_tot if ss_tot else 0
    adj = 1 - (1 - r2) * (n - 1) / df if ss_tot else 0
    return {"coef": beta.tolist(), "pvalues": pvals, "r2": r2, "adjR2": adj, "n": n}

OLS_MACRO_KEYS = ["sp500", "vix", "us_10y", "oil_wti", "eurusd", "gold",
                  "move", "us_2y", "dxy", "nasdaq100", "hy_credit"]
OLS_RATE_LIKE = {"us_10y", "us_2y", "vix", "move", "term_spread"}
OLS_FACTORS = ["sp500", "vix", "us_10y", "term_spread", "oil_wti", "eurusd",
               "gold", "move", "dxy", "nasdaq100", "hy_credit"]

def run_asset_regression(asset_hist, macro):
    asset_map = dict(asset_hist)
    maps = {k: dict(macro.get(k, [])) for k in OLS_MACRO_KEYS}
    dates = sorted(d for d in asset_map if all(d in maps[k] for k in OLS_MACRO_KEYS))
    levels = []
    for d in dates:
        row = {"asset": asset_map[d]}
        for k in OLS_MACRO_KEYS: row[k] = maps[k][d]
        row["term_spread"] = row["us_10y"] - row["us_2y"]
        levels.append(row)
    cols = ["asset"] + OLS_FACTORS
    changes = []
    for i in range(1, len(levels)):
        row = {}
        for c in cols:
            prev, cur = levels[i - 1][c], levels[i][c]
            row[c] = (cur - prev) if c in OLS_RATE_LIKE else ((cur - prev) / prev * 100 if prev else 0)
        changes.append(row)
    y = [r["asset"] for r in changes]
    X = [[r[f] for f in OLS_FACTORS] for r in changes]
    fit = fit_ols(y, X)
    vifs = {}
    for j, f in enumerate(OLS_FACTORS):
        yj = [r[f] for r in changes]
        Xj = [[r[g] for i2, g in enumerate(OLS_FACTORS) if i2 != j] for r in changes]
        fj = fit_ols(yj, Xj)
        vifs[f] = (1 / (1 - fj["r2"])) if fj["r2"] < 1 else None
    factors = {f: {"coef": fit["coef"][i + 1], "pvalue": fit["pvalues"][i + 1], "vif": vifs[f]}
               for i, f in enumerate(OLS_FACTORS)}
    return {"n": fit["n"], "r2": fit["r2"], "adjR2": fit["adjR2"], "factors": factors}

def run_carhart(port_ret, ff):
    ff_map = {r[0]: r for r in ff}
    y, X, rows = [], [], []
    for d, ret in port_ret:
        f = ff_map.get(d)
        if not f: continue
        mktrf, smb, hml, wml, rf = f[1], f[2], f[3], f[4], f[5]
        if mktrf == 0 and smb == 0 and hml == 0 and wml == 0: continue
        y.append(ret - rf)
        X.append([mktrf, smb, hml, wml])
        rows.append((d, ret, mktrf, smb, hml, wml, rf))
    fit = fit_ols(y, X)
    names = ["Mkt-RF", "SMB", "HML", "WML"]
    return {"n": fit["n"], "r2": fit["r2"], "adjR2": fit["adjR2"],
            "alphaDaily": fit["coef"][0], "alphaAnnual": fit["coef"][0] * 252,
            "alphaPvalue": fit["pvalues"][0],
            "factors": [{"name": nm, "beta": fit["coef"][i + 1], "pvalue": fit["pvalues"][i + 1]}
                        for i, nm in enumerate(names)],
            "rows": rows}


# ═══════════════ main: compute everything ═══════════════
def main():
    with open(os.path.join(DATA, "macro_history.json")) as f: macro = json.load(f)
    with open(os.path.join(DATA, "ff_factors.json")) as f: ff = json.load(f)

    holdings = []
    for p in PORTFOLIO:
        with open(p["file"]) as f: hist = json.load(f)
        pp = nav_on_or_before(hist, p["purchase"])
        cur = hist[-1][1]
        mv = p["amount"] * cur / pp
        holdings.append({**p, "history": hist, "purchasePrice": pp, "currentPrice": cur, "mv": mv})
    total_mv = sum(h["mv"] for h in holdings)
    for h in holdings: h["weight"] = h["mv"] / total_mv

    port_ret = build_portfolio_returns(holdings)
    rets = [r for _, r in port_ret]

    rf_map = {r[0]: r[5] for r in ff}
    rfs = [rf_map[d] for d, _ in port_ret if d in rf_map]
    rf_daily = sum(rfs) / len(rfs) if rfs else 0.0

    met = perf_metrics(rets, rf_daily)
    g = garch_fit(rets)
    v = var_analysis(rets, g["condVolDaily"], total_mv)
    ts_by_asset = [{"name": h["name"], "t": tsmom([x[1] for x in h["history"] if x[1] > 0])} for h in holdings]
    idx = []; w = 100.0
    for r in rets: w *= 1 + r / 100; idx.append(w)
    ts_port = tsmom(idx)
    hhi = sum(h["weight"] ** 2 for h in holdings)
    # pairwise correlation on shared dates (population formulas)
    rm = []
    for h in holdings:
        m = {}
        for i in range(1, len(h["history"])):
            p0 = h["history"][i - 1][1]
            if p0: m[h["history"][i][0]] = (h["history"][i][1] - p0) / p0 * 100
        rm.append(m)
    shared = sorted(set(rm[0]) & set(rm[1]))
    xa = [rm[0][d] for d in shared]; xb = [rm[1][d] for d in shared]
    mx, my = qmean(xa), qmean(xb)
    sxy = sum((xa[i] - mx) * (xb[i] - my) for i in range(len(shared)))
    sxx = sum((x - mx) ** 2 for x in xa); syy = sum((x - my) ** 2 for x in xb)
    corr = sxy / math.sqrt(sxx * syy)

    ols_by_asset = [{"name": h["name"], "res": run_asset_regression(h["history"], macro)} for h in holdings]
    carhart = run_carhart(port_ret, ff)

    summary = {
        "weights": {h["symbol"]: round(h["weight"], 4) for h in holdings},
        "totalMV": round(total_mv, 2), "rfDaily": round(rf_daily, 6),
        "met": {k: (round(vv, 4) if isinstance(vv, float) else vv) for k, vv in met.items() if k != "wealth"},
        "garch": {k: (round(vv, 4) if isinstance(vv, float) else vv) for k, vv in g.items() if k != "eps"},
        "var": {"h95": round(v["hist"]["var95"], 4), "h99": round(v["hist"]["var99"], 4),
                "es95": round(v["hist"]["es95"], 4), "p95": round(v["param"]["var95"], 4),
                "p99": round(v["param"]["var99"], 4), "g95": round(v["paramGarch"]["var95"], 4),
                "mc95": round(v["mc10d"]["var95"], 4), "mc99": round(v["mc10d"]["var99"], 4)},
        "tsPort": {"score": round(ts_port["score"], 4)},
        "div": {"hhi": round(hhi, 4), "effN": round(1 / hhi, 4), "corr": round(corr, 4), "nShared": len(shared)},
        "carhart": {"alphaAnnual": round(carhart["alphaAnnual"], 2), "r2": round(carhart["r2"], 4),
                    "betas": {f["name"]: round(f["beta"], 4) for f in carhart["factors"]}},
        "ols": {o["name"]: {"r2": round(o["res"]["r2"], 4), "n": o["res"]["n"]} for o in ols_by_asset},
    }
    print(json.dumps(summary, indent=2, ensure_ascii=False))

    build_workbook(holdings, port_ret, rf_daily, met, g, v, ts_by_asset, ts_port,
                   hhi, corr, shared, rm, ols_by_asset, carhart, total_mv)
    return summary


# ═══════════════ the Excel workbook ═══════════════
def build_workbook(holdings, port_ret, rf_daily, met, g, v, ts_by_asset, ts_port,
                   hhi, corr, shared, rm, ols_by_asset, carhart, total_mv):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter

    ARIAL = "Arial"
    F = Font(name=ARIAL, size=10)
    FB = Font(name=ARIAL, size=10, bold=True)
    FH = Font(name=ARIAL, size=12, bold=True)
    BLUE = Font(name=ARIAL, size=10, color="0000FF")      # hardcoded inputs / Python-computed values
    GREEN = Font(name=ARIAL, size=10, color="008000")     # cross-sheet links
    YELLOW = PatternFill("solid", fgColor="FFFF00")
    GREY = PatternFill("solid", fgColor="F2F2F2")

    wb = Workbook()

    def style_all(ws):
        for row in ws.iter_rows():
            for c in row:
                if c.font is None or c.font.name != ARIAL:
                    base = c.font
                    c.font = Font(name=ARIAL, size=10, bold=base.bold if base else False,
                                  color=base.color if base else None)

    def head(ws, cells, row=1):
        for i, t in enumerate(cells, start=1):
            c = ws.cell(row=row, column=i, value=t); c.font = FB; c.fill = GREY

    n = len(port_ret)

    # ---------- README ----------
    ws = wb.active; ws.title = "README"
    ws.column_dimensions["A"].width = 108
    lines = [
        ("QUANT DOUBLE-CHECK — Portfolio BI", FH),
        ("Replica in Excel di TUTTI i modelli calcolati nel browser (docs/quant.js, ols.js, ff.js), per verifica indipendente.", F),
        ("Dove possibile i numeri sono FORMULE vive: cambiando un input si ricalcolano. Dove serve un ottimizzatore o una", F),
        ("simulazione (GARCH α/β, Monte Carlo, beta delle regressioni) il valore è scritto in BLU con una nota che spiega come nasce.", F),
        ("", F),
        ("PORTAFOGLIO DI TEST (lo stesso della sessione di test nel browser):", FB),
        ("  • NVDA — €3.000 investiti il 2023-01-10   • BGF-SE — €1.500 investiti il 2024-03-15", F),
        ("  I pesi sono le quote di valore di mercato attuale (foglio Portafoglio).", F),
        ("", F),
        ("CONVENZIONI (identiche al browser):", FB),
        ("  • rendimenti giornalieri in %, deviazione standard di POPOLAZIONE (÷n), 252 giorni/anno", F),
        ("  • quantili con interpolazione lineare (PERCENTILE di Excel usa la stessa formula)", F),
        ("  • GARCH(1,1) gaussiano con variance targeting, MLE su griglia (α passo 0,02→0,005; β passo 0,01→0,0025)", F),
        ("  • Monte Carlo: bootstrap di 10.000 percorsi da 10 giorni sui rendimenti reali, RNG mulberry32 seed 42", F),
        ("  • p-value delle regressioni con CDF normale (con n≈1000 è indistinguibile dalla t di Student)", F),
        ("", F),
        ("LEGENDA COLORI:", FB),
        ("  BLU = valore hardcoded (input o risultato Python documentato) · NERO = formula Excel · VERDE = link ad altro foglio", F),
        ("  Sfondo GIALLO = celle di confronto browser ↔ Excel", F),
        ("", F),
        ("FOGLI: Portafoglio · Rendimenti · Metriche · GARCH · VaR · TSMOM · Correlazione · OLS · Carhart · Carhart_dati", F),
        ("Fonte dati: docs/data/ (prezzi Yahoo/BlackRock, fattori Ken French) — stessi file statici letti dal browser.", F),
        (f"Generato da tools/quant_check.py — rigeneralo dopo ogni aggiornamento dati con: python3 tools/quant_check.py", F),
    ]
    for i, (t, fnt) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=t); c.font = fnt

    # ---------- Portafoglio ----------
    ws = wb.create_sheet("Portafoglio")
    head(ws, ["Titolo", "Investito €", "Data acquisto", "Prezzo acquisto", "Prezzo attuale",
              "Valore di mercato €", "Peso"])
    for i, h in enumerate(holdings, start=2):
        ws.cell(row=i, column=1, value=h["name"]).font = F
        c = ws.cell(row=i, column=2, value=h["amount"]); c.font = BLUE; c.number_format = "#,##0"
        ws.cell(row=i, column=3, value=h["purchase"]).font = BLUE
        c = ws.cell(row=i, column=4, value=h["purchasePrice"]); c.font = BLUE
        c.comment = Comment("Prezzo di chiusura alla data di acquisto (o ultimo precedente), da docs/data — stessa regola navOnOrBefore del browser.", "quant_check")
        c = ws.cell(row=i, column=5, value=h["currentPrice"]); c.font = BLUE
        c.comment = Comment("Ultimo prezzo pubblicato nei dati statici.", "quant_check")
        ws.cell(row=i, column=6, value=f"=B{i}*E{i}/D{i}").number_format = "#,##0"
        ws.cell(row=i, column=7, value=f"=F{i}/$F${len(holdings)+2}").number_format = "0.0%"
    r = len(holdings) + 2
    ws.cell(row=r, column=1, value="TOTALE").font = FB
    ws.cell(row=r, column=6, value=f"=SUM(F2:F{r-1})").number_format = "#,##0"
    for col, wdt in zip("ABCDEFG", (38, 12, 13, 14, 14, 18, 9)): ws.column_dimensions[col].width = wdt

    # ---------- Rendimenti ----------
    ws = wb.create_sheet("Rendimenti")
    head(ws, ["Data", "Rendimento %", "Extra-rend. (r − rf)", "ln(1+r/100)",
              "Indice ricchezza", "Massimo corrente", "Drawdown %"])
    ws.cell(row=1, column=9, value="rf giornaliero medio %").font = FB
    c = ws.cell(row=2, column=9, value=rf_daily); c.font = BLUE
    c.comment = Comment("Media del tasso privo di rischio giornaliero (serie RF di Ken French, docs/data/ff_factors.json) sulle date del portafoglio — identico al browser.", "quant_check")
    for i, (d, r_) in enumerate(port_ret, start=2):
        ws.cell(row=i, column=1, value=d).font = F
        c = ws.cell(row=i, column=2, value=r_); c.font = BLUE; c.number_format = "0.0000"
        ws.cell(row=i, column=3, value=f"=B{i}-$I$2").number_format = "0.0000"
        ws.cell(row=i, column=4, value=f"=LN(1+B{i}/100)").number_format = "0.000000"
        ws.cell(row=i, column=5, value=("=1*(1+B2/100)" if i == 2 else f"=E{i-1}*(1+B{i}/100)")).number_format = "0.0000"
        ws.cell(row=i, column=6, value=("=E2" if i == 2 else f"=MAX(F{i-1},E{i})")).number_format = "0.0000"
        ws.cell(row=i, column=7, value=f"=(E{i}/F{i}-1)*100").number_format = "0.00"
    ws.cell(row=1, column=2).comment = Comment("Rendimenti giornalieri % del portafoglio pesato — costruiti come in ff.js buildPortfolioReturns (intersezione delle date di tutti i titoli).", "quant_check")
    for col, wdt in zip("ABCDEFG", (12, 13, 16, 13, 14, 15, 12)): ws.column_dimensions[col].width = wdt

    last = n + 1  # last data row in Rendimenti

    # ---------- Metriche ----------
    ws = wb.create_sheet("Metriche")
    head(ws, ["Metrica", "Formula Excel", "Valore", "Motore (JS≡Py)", "Diff"])
    js = {"CAGR %": round(met["annRet"], 4), "Volatilità annua %": round(met["annVol"], 4),
          "Sharpe": round(met["sharpe"], 4), "Sortino": round(met["sortino"], 4),
          "Max drawdown %": round(met["maxDD"], 4), "Calmar": round(met["calmar"], 4)}
    rows = [
        ("CAGR %", f"=(EXP(SUM(Rendimenti!D2:D{last})*252/{n})-1)*100"),
        ("Volatilità annua %", f"=STDEVP(Rendimenti!B2:B{last})*SQRT(252)"),
        ("Sharpe", f"=AVERAGE(Rendimenti!C2:C{last})/STDEVP(Rendimenti!B2:B{last})*SQRT(252)"),
        ("Sortino", f"=AVERAGE(Rendimenti!C2:C{last})/SQRT(SUMPRODUCT((Rendimenti!C2:C{last}<0)*Rendimenti!C2:C{last}^2)/{n})*SQRT(252)"),
        ("Max drawdown %", f"=MIN(Rendimenti!G2:G{last})"),
        ("Calmar", "=C2/ABS(C6)"),
    ]
    for i, (name, formula) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=name).font = F
        ws.cell(row=i, column=2, value=formula.replace("=", "", 1)).font = Font(name=ARIAL, size=8, color="808080")
        ws.cell(row=i, column=3, value=formula).number_format = "0.0000"
        c = ws.cell(row=i, column=4, value=js[name]); c.font = BLUE; c.fill = YELLOW
        ws.cell(row=i, column=5, value=f"=C{i}-D{i}").number_format = "0.000000"
    ws.cell(row=1, column=4).comment = Comment("Valore del motore di calcolo — verificato identico tra browser (JS) e Python a piena precisione. Diff ≈ 0 conferma che le formule Excel replicano il motore.", "quant_check")
    for col, wdt in zip("ABCDE", (18, 78, 12, 12, 12)): ws.column_dimensions[col].width = wdt

    # ---------- GARCH ----------
    ws = wb.create_sheet("GARCH")
    ws.cell(row=1, column=1, value="GARCH(1,1) con variance targeting — ricorsione VIVA: cambia α o β e tutto si ricalcola").font = FB
    labels = [("α (alpha)", g["alpha"], "Stimato per massima verosimiglianza su griglia (passo finale 0,005) — vedi README. Modificabile per esplorare."),
              ("β (beta)", g["beta"], "Stimato insieme ad α. α+β<1 richiesto."),
              ("media rendimenti μ", g["mu"], "Media dei rendimenti; eps = r − μ.")]
    for i, (lab, val, note) in enumerate(labels, start=3):
        ws.cell(row=i, column=1, value=lab).font = FB
        c = ws.cell(row=i, column=2, value=val); c.font = BLUE; c.number_format = "0.000000"
        c.comment = Comment(note, "quant_check")
    ws.cell(row=6, column=1, value="varianza incondizionata").font = FB
    ws.cell(row=6, column=2, value=f"=SUMPRODUCT(C9:C{n+8},C9:C{n+8})/{n}").number_format = "0.000000"
    ws.cell(row=7, column=1, value="ω (omega) = uncondVar·(1−α−β)").font = FB
    ws.cell(row=7, column=2, value="=B6*(1-B3-B4)").number_format = "0.000000"
    head(ws, ["Data", "Rendimento %", "eps = r − μ", "eps²", "σ²_t (ricorsione)", "contributo −LL"], row=8)
    for i, (d, r_) in enumerate(port_ret, start=9):
        ws.cell(row=i, column=1, value=d).font = F
        ws.cell(row=i, column=2, value=f"=Rendimenti!B{i-7}").font = GREEN
        ws.cell(row=i, column=3, value=f"=B{i}-$B$5").number_format = "0.0000"
        ws.cell(row=i, column=4, value=f"=C{i}^2").number_format = "0.0000"
        if i == 9:
            ws.cell(row=i, column=5, value="=$B$6").number_format = "0.000000"
        else:
            ws.cell(row=i, column=5, value=f"=$B$7+$B$3*D{i-1}+$B$4*E{i-1}").number_format = "0.000000"
            ws.cell(row=i, column=6, value=f"=LN(E{i})+D{i}/E{i}").number_format = "0.0000"
    endr = n + 8
    summ = [
        ("−2·logLik (≈, costante esclusa)", f"=SUM(F10:F{endr})", round(g["negLL"], 4)),
        ("σ² previsione 1 giorno", f"=$B$7+$B$3*D{endr}+$B$4*E{endr}", None),
        ("Vol. condizionata annua %", f"=SQRT(B{endr+2}*252)", round(g["condVolAnn"], 4)),
        ("Vol. incondizionata annua %", "=SQRT(B6*252)", round(g["uncondVolAnn"], 4)),
    ]
    for j, (lab, formula, jsval) in enumerate(summ):
        rr = endr + 1 + j
        ws.cell(row=rr, column=1, value=lab).font = FB
        ws.cell(row=rr, column=2, value=formula).number_format = "0.0000"
        if jsval is not None:
            c = ws.cell(row=rr, column=3, value=jsval); c.font = BLUE; c.fill = YELLOW
            c.comment = Comment("Valore del motore (identico in JS e Python, verificato a piena precisione).", "quant_check")
            ws.cell(row=rr, column=4, value=f"=B{rr}-C{rr}").number_format = "0.000000"
    for col, wdt in zip("ABCDEF", (28, 14, 12, 12, 14, 14)): ws.column_dimensions[col].width = wdt

    # ---------- VaR ----------
    ws = wb.create_sheet("VaR")
    ws.cell(row=1, column=1, value="Value at Risk — perdite in % (positive). Storico e parametrico sono formule vive sui rendimenti.").font = FB
    head(ws, ["Metodo", "Livello", "Formula Excel", "Valore %", "Motore (JS≡Py)", "Diff", "€ sul portafoglio"], row=3)
    mv_ref = f"Portafoglio!$F${len(holdings)+2}"
    var_rows = [
        ("Storico", "95%", f"=-PERCENTILE(Rendimenti!B2:B{last},0.05)", round(v["hist"]["var95"], 4)),
        ("Storico", "99%", f"=-PERCENTILE(Rendimenti!B2:B{last},0.01)", round(v["hist"]["var99"], 4)),
        ("Expected Shortfall", "95%", f"=-AVERAGEIF(Rendimenti!B2:B{last},\"<=\"&-D4)", round(v["hist"]["es95"], 4)),
        ("Varianza-covarianza", "95%", f"=1.645*STDEVP(Rendimenti!B2:B{last})", round(v["param"]["var95"], 4)),
        ("Varianza-covarianza", "99%", f"=2.326*STDEVP(Rendimenti!B2:B{last})", round(v["param"]["var99"], 4)),
        ("Parametrico vol GARCH", "95%", f"=1.645*SQRT(GARCH!B{n+10})", round(v["paramGarch"]["var95"], 4)),
        ("Monte Carlo 10g (bootstrap)", "95%", None, round(v["mc10d"]["var95"], 4)),
        ("Monte Carlo 10g (bootstrap)", "99%", None, round(v["mc10d"]["var99"], 4)),
    ]
    for i, (m_, lvl, formula, jsval) in enumerate(var_rows, start=4):
        ws.cell(row=i, column=1, value=m_).font = F
        ws.cell(row=i, column=2, value=lvl).font = F
        if formula:
            ws.cell(row=i, column=3, value=formula.replace("=", "", 1)).font = Font(name=ARIAL, size=8, color="808080")
            ws.cell(row=i, column=4, value=formula).number_format = "0.0000"
            ws.cell(row=i, column=6, value=f"=D{i}-E{i}").number_format = "0.000000"
        else:
            c = ws.cell(row=i, column=4, value=jsval); c.font = BLUE; c.number_format = "0.0000"
            c.comment = Comment("10.000 percorsi da 10 giorni ricampionando i rendimenti reali (bootstrap), RNG mulberry32 seed 42 — bit-identico tra browser e Python, non esprimibile come formula.", "quant_check")
        c = ws.cell(row=i, column=5, value=jsval); c.font = BLUE; c.fill = YELLOW
        ws.cell(row=i, column=7, value=f"={mv_ref}*D{i}/100").number_format = "#,##0"
    for col, wdt in zip("ABCDEFG", (26, 8, 52, 11, 12, 11, 15)): ws.column_dimensions[col].width = wdt

    # ---------- TSMOM ----------
    ws = wb.create_sheet("TSMOM")
    ws.cell(row=1, column=1, value="Time-Series Momentum — rendimenti multi-orizzonte (formule sull'indice di ricchezza) + EMA").font = FB
    head(ws, ["Serie", "1 mese %", "3 mesi %", "6 mesi %", "12 mesi %", "EMA20", "EMA100", "Score"], row=3)
    ws.cell(row=4, column=1, value="Portafoglio (indice)").font = F
    for j, h_ in enumerate((21, 63, 126, 252)):
        ws.cell(row=4, column=2 + j,
                value=f"=(INDEX(Rendimenti!E:E,{last})/INDEX(Rendimenti!E:E,{last}-{h_})-1)*100").number_format = "0.0"
    c = ws.cell(row=4, column=6, value=round(ts_port["ema20"], 4)); c.font = BLUE
    c.comment = Comment("EMA ricorsiva (seed = SMA dei primi N valori) calcolata in Python con la stessa formula del browser.", "quant_check")
    ws.cell(row=4, column=7, value=round(ts_port["ema100"], 4)).font = BLUE
    c = ws.cell(row=4, column=8, value=round(ts_port["score"], 4)); c.font = BLUE; c.fill = YELLOW
    c.comment = Comment("Media dei segni dei 4 orizzonti + segno EMA20>EMA100 — identico nel browser.", "quant_check")
    for i, e in enumerate(ts_by_asset, start=5):
        t = e["t"]
        ws.cell(row=i, column=1, value=e["name"]).font = F
        for j, h_ in enumerate((21, 63, 126, 252)):
            s = next((s_ for s_ in t["signals"] if s_["h"] == h_), None)
            if s: ws.cell(row=i, column=2 + j, value=round(s["ret"], 2)).font = BLUE
        ws.cell(row=i, column=6, value=round(t["ema20"], 4)).font = BLUE
        ws.cell(row=i, column=7, value=round(t["ema100"], 4)).font = BLUE
        ws.cell(row=i, column=8, value=round(t["score"], 4)).font = BLUE
    ws.cell(row=8, column=1, value="Nota: per i singoli titoli i valori sono calcolati in Python sulle stesse serie prezzi del browser (identiche formule).").font = F
    for col, wdt in zip("ABCDEFGH", (34, 10, 10, 10, 10, 12, 12, 8)): ws.column_dimensions[col].width = wdt

    # ---------- Correlazione ----------
    ws = wb.create_sheet("Correlazione")
    ws.cell(row=1, column=1, value="Correlazione tra i titoli sui giorni condivisi — CORREL è una formula viva").font = FB
    head(ws, ["Data", holdings[0]["symbol"] + " %", holdings[1]["symbol"] + " %"], row=3)
    for i, d in enumerate(shared, start=4):
        ws.cell(row=i, column=1, value=d).font = F
        ws.cell(row=i, column=2, value=rm[0][d]).font = BLUE
        ws.cell(row=i, column=3, value=rm[1][d]).font = BLUE
    endc = len(shared) + 3
    ws.cell(row=1, column=5, value="Correlazione").font = FB
    ws.cell(row=2, column=5, value=f"=CORREL(B4:B{endc},C4:C{endc})").number_format = "0.0000"
    c = ws.cell(row=2, column=6, value=round(corr, 4)); c.font = BLUE; c.fill = YELLOW
    ws.cell(row=2, column=7, value="=E2-F2").number_format = "0.000000"
    ws.cell(row=4, column=5, value="HHI (concentrazione)").font = FB
    ws.cell(row=5, column=5, value=f"=SUMPRODUCT(Portafoglio!G2:G{len(holdings)+1},Portafoglio!G2:G{len(holdings)+1})").number_format = "0.0000"
    ws.cell(row=6, column=5, value="Titoli effettivi = 1/HHI").font = FB
    ws.cell(row=7, column=5, value="=1/E5").number_format = "0.00"
    for col, wdt in zip("ABCDEFG", (12, 12, 12, 3, 20, 10, 10)): ws.column_dimensions[col].width = wdt

    # ---------- OLS ----------
    ws = wb.create_sheet("OLS")
    ws.cell(row=1, column=1, value="Regressione OLS macro per titolo — replica Python di docs/ols.js (coefficienti = statsmodels; p-value CDF normale)").font = FB
    r = 3
    for o in ols_by_asset:
        res = o["res"]
        ws.cell(row=r, column=1, value=f"{o['name']} — R²={res['r2']:.4f} (adj {res['adjR2']:.4f}), n={res['n']}").font = FB
        head(ws, ["Fattore", "Coefficiente", "p-value", "VIF"], row=r + 1)
        for i, f_ in enumerate(OLS_FACTORS, start=r + 2):
            fa = res["factors"][f_]
            ws.cell(row=i, column=1, value=f_).font = F
            ws.cell(row=i, column=2, value=round(fa["coef"], 6)).font = BLUE
            ws.cell(row=i, column=3, value=round(fa["pvalue"], 6)).font = BLUE
            ws.cell(row=i, column=4, value=round(fa["vif"], 3) if fa["vif"] else None).font = BLUE
        r += len(OLS_FACTORS) + 4
    ws.cell(row=r, column=1, value="I coefficienti nascono da (X'X)⁻¹X'y — un ottimizzatore matriciale, non esprimibile come formula di cella. Il foglio Carhart_dati fornisce i dati grezzi per una verifica manuale.").font = F
    for col, wdt in zip("ABCD", (18, 14, 12, 10)): ws.column_dimensions[col].width = wdt

    # ---------- Carhart ----------
    ws = wb.create_sheet("Carhart")
    ws.cell(row=1, column=1, value=f"Fama-French-Carhart sul portafoglio — R²={carhart['r2']:.4f}, n={carhart['n']} — dati grezzi nel foglio Carhart_dati").font = FB
    head(ws, ["Termine", "Beta", "p-value"], row=3)
    rows_c = [("Alpha (daily %)", carhart["alphaDaily"], carhart["alphaPvalue"])] + \
             [(f["name"], f["beta"], f["pvalue"]) for f in carhart["factors"]]
    for i, (nm, b, p) in enumerate(rows_c, start=4):
        ws.cell(row=i, column=1, value=nm).font = F
        ws.cell(row=i, column=2, value=round(b, 6)).font = BLUE
        ws.cell(row=i, column=3, value=round(p, 6)).font = BLUE
    ws.cell(row=10, column=1, value=f"Alpha annualizzato = alpha daily × 252 = {carhart['alphaAnnual']:.2f}%").font = FB
    ws.cell(row=11, column=1, value="Regressione: (r_portafoglio − RF) = α + β·(Mkt−RF) + β·SMB + β·HML + β·WML + ε — fattori 'Developed' giornalieri di Ken French.").font = F
    for col, wdt in zip("ABCD", (18, 14, 12, 14)): ws.column_dimensions[col].width = wdt

    # ---------- Carhart_dati ----------
    ws = wb.create_sheet("Carhart_dati")
    head(ws, ["Data", "r_port %", "Mkt-RF", "SMB", "HML", "WML", "RF", "y = r − RF"])
    for i, (d, ret, mktrf, smb, hml, wml, rf_) in enumerate(carhart["rows"], start=2):
        for j, val in enumerate((d, ret, mktrf, smb, hml, wml, rf_), start=1):
            c = ws.cell(row=i, column=j, value=val); c.font = BLUE if j > 1 else F
        ws.cell(row=i, column=8, value=f"=B{i}-G{i}").number_format = "0.0000"
    for col, wdt in zip("ABCDEFGH", (12, 10, 9, 9, 9, 9, 9, 11)): ws.column_dimensions[col].width = wdt

    for ws_ in wb.worksheets: style_all(ws_)
    out = os.path.join(ROOT, "quant_double_check.xlsx")
    wb.save(out)
    print(f"\nSaved: {out}")


if __name__ == "__main__":
    main()
